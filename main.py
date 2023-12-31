import codecs
import glob
import json
import os
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from telnetlib import EC

import openpyxl
import pandas as pd
from appium import webdriver
from appium.options.common.base import AppiumOptions
from appium.webdriver.common.appiumby import AppiumBy
from openpyxl.styles import Alignment
from PyQt5 import QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QMouseEvent, QPalette
from PyQt5.QtWidgets import (QApplication, QCheckBox, QComboBox, QDialog,
                             QHBoxLayout, QHeaderView, QLabel, QLineEdit,
                             QMessageBox, QPushButton, QTableWidget,
                             QTableWidgetItem, QVBoxLayout, QWidget)
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

df = pd.read_excel('du_lieu/data.xlsx')
customer_code = df.iloc[:, 0].dropna().values.tolist()


def get_devices():
    devices = {}
    for device in subprocess.run(["adb", "devices", "-l"], capture_output=True).stdout.decode().split("\n")[1:-2]:
        if device.strip():
            device_adb_name = device.split()[0].strip()
            device_name = subprocess.run(
                ["adb", "-s", f"{device_adb_name}", "shell", "getprop", "ro.product.model"], capture_output=True).stdout.decode().strip()
            if device_name:
                devices[device_name] = device_adb_name
    return devices


devices = get_devices()


def select_devices(selected_device_names):
    selected_device_adb_names = []
    for selected_device_name in selected_device_names:
        if selected_device_name in devices:
            selected_device_adb_names.append(devices[selected_device_name])
        else:
            return "Vui lòng khởi động lại để quét thiết bị Offline"
    return selected_device_adb_names

SCREENSHOT_PATH = os.path.join(os.getcwd(), "images")
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BillEaseBank")
        self.layout = QVBoxLayout()
        self.resize(1024, 576)
        self.setWindowFlags(QtCore.Qt.WindowType.WindowMaximizeButtonHint |
                            QtCore.Qt.WindowType.WindowMinimizeButtonHint | QtCore.Qt.WindowType.WindowCloseButtonHint)
        palette = QPalette()
        self.selected_devices = []
        self.screenshot = 0
        # username, ma_khach_hang, userpayment, self.screenshot
        self.username = ""
        self.ma_khach_hang = ""
        self.userpayment = ""
        palette.setColor(QPalette.Window, QColor("#282a36"))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor("#282a36"))
        palette.setColor(QPalette.AlternateBase, QColor("#353746"))
        palette.setColor(QPalette.ToolTipBase, QColor("#f8f8f2"))
        palette.setColor(QPalette.ToolTipText, QColor("#f8f8f2"))
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor("#6272a4"))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Link, QColor("#ff79c6"))
        palette.setColor(QPalette.Highlight, QColor("#bd93f9"))
        palette.setColor(QPalette.HighlightedText, QColor("#282a36"))
        self.setPalette(palette)

        self.scan_button = QPushButton("Quét thiết bị")
        self.scan_button.setStyleSheet("background-color: #FF79C6")
        self.start_button = QPushButton("Chạy")
        self.start_button.setStyleSheet("background-color: #50FA7B")
        self.start_button.setCursor(Qt.PointingHandCursor)
        self.stop_button = QPushButton("Dừng")
        self.refresh_button = QPushButton("Refresh")
        self.screenshot_checkbox = QCheckBox("Chụp màn hình")
        self.username_label = QLabel("Tên tài khoản")
        self.username_label.setStyleSheet("color: #8BE9FD;")
        self.username_label.setStyleSheet("border: none;")
        self.username_input = QLineEdit()
        self.userpayment_input = QLineEdit()
        self.bank_combobox = QComboBox()

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderItem(0, QTableWidgetItem("STT"))
        self.table.setHorizontalHeaderItem(
            1, QTableWidgetItem("Mã khách hàng"))
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setShowGrid(True)
        self.table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarAlwaysOff)
        self.table.setColumnWidth(0, 60)
        self.table.setColumnWidth(1, 140)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.scan_button)
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.stop_button)
        button_layout.addWidget(self.refresh_button)

        self.layout.addLayout(button_layout)
        self.layout.addWidget(self.username_label, alignment=Qt.AlignCenter)
        self.layout.addWidget(self.username_input)
        self.layout.addWidget(self.userpayment_input)
        self.layout.addWidget(self.bank_combobox)
        self.layout.addWidget(self.screenshot_checkbox)
        self.layout.addWidget(self.table)

        self.setLayout(self.layout)
        self.scan_button.clicked.connect(self.scan_devices)
        self.start_button.clicked.connect(self.start_process)
        self.stop_button.clicked.connect(self.stop_process)
        self.refresh_button.clicked.connect(self.update_table)
        self.screenshot_checkbox.stateChanged.connect(self.toggle_screenshot)
        try:
            with codecs.open('permission/banks.json', 'r', 'utf-8-sig') as f:
                banks = json.load(f)
        except:
            quit()
        self.bank_combobox.addItem("Chọn ngân hàng")
        self.bank_combobox.addItems(banks)
        self.fill_default_values()
        self.update_table()
        # self.showMaximized()
        self.table.setStyleSheet("""
            QTableWidget::item {
                padding-top: 3px;
                padding-bottom: 3px;
            }
        """)

        self.setStyleSheet("""
            * {
                font-family: 'JetBrains Mono', monospace;
                color: #f8f8f2;
                background-color: #282a36;
                font-size: 16px;
            }
            QPushButton, QLineEdit, QCheckBox, QTableWidget, QHeaderView::section {
                font-size: 16px;
            }
            QPushButton, QLineEdit {
                background-color: #44475A;
                color: #f8f8f2;
                padding: 15px;
                text-align: center;
                text-decoration: none;
                margin: 4px 2px;
                border-radius: 20px;
            }
            QLineEdit {
                padding: 10px;
                border-radius: 0px;
            }
            QPushButton {
                background-color: #6272a4;
            }
            QPushButton:hover {
                background-color: #44475a;
            }
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
                border-radius: 15px;
            }
            QCheckBox::indicator:unchecked {
                background-color: #44475a;
                border: 1px solid #bd93f9;
            }
            QCheckBox::indicator:checked {
                background-color: #bd93f9;
            }
            QTableWidget {
                background-color: #282a36;
                color: #f8f8f2;
                gridline-color: #6272a4;
                border-top-left-radius: 0px;
                border-top-right-radius: 0px;
                border-bottom-left-radius: 20px;
                border-bottom-right-radius: 20px;
                border: 1px solid #6272a4;
            }
            QHeaderView::section {
                background-color: #282a36;
                color: white;
                padding: 4px;
                border: none;
            }
            QTableView {
                selection-background-color: #6272A4;
                selection-color: #F8F8F2;
            }
            QTableView::item:selected {
                background-color: #6272A4;
                color: #F8F8F2;
                border: none;
            }
        """)
        self.bank_combobox.setStyleSheet("""
            QComboBox {
                background-color: #44475A;
                font-size: 16px;
                border: none;
                padding: 10px;
                color: #f8f8f2;
            }
            QComboBox::drop-down {
                background-color: #44475A;
                border-radius: 20px;
            }
        """)

    def update_table(self):
        try:
            df = pd.read_excel('du_lieu/data.xlsx')
            customer_code = df.iloc[:, 0].dropna().values.tolist()
            row_count = len(customer_code)
            self.table.setRowCount(row_count)
            index = 0
            for row in range(row_count):
                index += 1
                self.table.setItem(row, 0, QTableWidgetItem(str(index)))
                self.table.setItem(
                    row, 1, QTableWidgetItem(customer_code[row]))
                self.table.item(row, 0).setTextAlignment(Qt.AlignCenter)
                self.table.item(row, 1).setTextAlignment(Qt.AlignCenter)
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Error occurred while updating table: {str(e)}")

    def fill_default_values(self):
        if self.username_input.text().strip() == "":
            self.username_input.setPlaceholderText("Điền tên thẻ thanh toán")
            self.userpayment_input.setPlaceholderText("Điền tên người thanh toán")
            self.bank_combobox.setCurrentText("Chọn ngân hàng")

    def scan_devices(self):
        popup = ScanPopup()
        popup.exec_()

    def start_process(self):
        self.start_button.setEnabled(False)
        self.start_button.setStyleSheet("background-color: #bd93f9")
        self.stop_button.setEnabled(True)
        bank = self.bank_combobox.currentText()
        self.username = self.username_input.text()
        self.userpayment = self.userpayment_input.text()
        df = pd.read_excel('du_lieu/data.xlsx')
        self.ma_khach_hang = df.iloc[:, 0].dropna().values.tolist()
        if bank == "VPBank":
            # self.vpbank(username, ma_khach_hang, userpayment, self.screenshot)
            vpbank_thread = threading.Thread(target=self.vpbank)
            vpbank_thread.start()
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
        else:
            error = Error("Vui chọn ngân hàng!")
            error.exec_()
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
            pass
    def stop_process(self):
        pass

    def toggle_screenshot(self, state):
        self.screenshot = state


    def vpbank(self):
        df = pd.DataFrame()
        SCREENSHOT_PATH = os.path.join(os.getcwd(), "images")
        account_name = self.username
        ma_kh = self.ma_khach_hang
        screen_shot = False
        if self.screenshot == 2:
            screen_shot = True
        global error
        error = ["Không có lỗi"]

        def run(device, account_name, ma_kh, screenshot=False):
            global tien_number
            global error
            options = AppiumOptions()
            options.load_capabilities({
                "deviceName": f"{device}",
                "platformName": "Android",
                "automationName": "UiAutomator2",
                "udid": f"{device}"
            })
            try:
                driver = webdriver.Remote(
                    "http://127.0.0.1:4723", options=options)
            except:
                driver = webdriver.Remote(
                    "http://127.0.0.1:4723", options=options)
            wait = WebDriverWait(driver, 5)
            try:
                thanh_toan = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
                thanh_toan.click()
            except:
                pass
            global count_kh
            count_kh = 0
            kh = ma_kh[count_kh]
            total_wait_time = 0
            s_time = time.time()
            try:
                dien = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
                dien.click()
            except StaleElementReferenceException:
                dien = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
                dien.click()
            except:
                pass
            end_time = time.time() - s_time
            for_total_time = total_wait_time - end_time
            for i in ma_kh:
                if i in error:
                    continue
                num_steps = 5
                start_time = time.time()

                try:
                    new_bill = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/ll_create_new_bill"]')))
                    new_bill.click()
                except:
                    pass
                nha_cungcap = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/llValue"]')))
                nha_cungcap.click()
                elapsed_time = time.time() - start_time
                if elapsed_time < for_total_time:
                    remaining_wait_time = for_total_time - elapsed_time
                else:
                    remaining_wait_time = 0

                if remaining_wait_time > 0:
                    wait_time_per_step = remaining_wait_time / (num_steps - 1)
                    time.sleep(wait_time_per_step)
                toan_quoc = wait.until(EC.presence_of_element_located((
                    AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvName" and @text="Điện lực toàn quốc"]'
                )))
                toan_quoc.click()
                try:
                    wait_tai_khoan_nguon = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='com.vnpay.vpbankonline:id/selectAccount']/android.widget.LinearLayout/android.widget.LinearLayout")))
                    wait_tai_khoan_nguon.click()
                except StaleElementReferenceException:
                    wait_tai_khoan_nguon = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='com.vnpay.vpbankonline:id/selectAccount']/android.widget.LinearLayout/android.widget.LinearLayout")))
                    wait_tai_khoan_nguon.click()
                except:
                    error = "Lỗi không xác định"
                    return error

                def chon_tai_khoan_nguon(account):
                    valid_input = False
                    while not valid_input:
                        elements = driver.find_element(by=AppiumBy.XPATH, value="//android.widget.ListView[@resource-id='com.vnpay.vpbankonline:id/listAccount']") \
                            .find_elements(by=AppiumBy.CLASS_NAME, value="android.widget.TextView")
                        for element in elements:
                            if element.text == account:
                                valid_input = True
                                return element
                        return "Chỉ cần nhập số thẻ hoặc loại thẻ"

                chon_tai_khoan_nguon(account_name).click()
                elapsed_time = time.time() - start_time
                if elapsed_time < for_total_time:
                    remaining_wait_time = for_total_time - elapsed_time
                else:
                    remaining_wait_time = 0

                if remaining_wait_time > 0:
                    wait_time_per_step = remaining_wait_time / (num_steps - 2)
                    time.sleep(wait_time_per_step)

                def input(kh):
                    global count_kh
                    input_mkh = wait.until(EC.presence_of_element_located((
                        AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.vnpay.vpbankonline:id/edtCustomerCode"]')))
                    input_mkh.send_keys(kh)
                    tiep = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTitleAction"]')))
                    tiep.click()
                    try:
                        error_text = wait.until(EC.presence_of_element_located(
                            (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvError"]'))
                        ).get_attribute('text')
                        clear = wait.until(EC.presence_of_element_located(
                            (AppiumBy.XPATH, '//android.widget.ImageView[@resource-id="com.vnpay.vpbankonline:id/vClear"]')))
                        clear.click()
                        count_kh += 1
                        ma_kh[count_kh - 1] = kh + f': {error_text}'
                        if count_kh == len(ma_kh):
                            return error
                        kh = ma_kh[count_kh]
                        input(kh)
                    except:
                        pass
                    return ma_kh
                error = input(kh)

                elapsed_time = time.time() - start_time
                if elapsed_time < for_total_time:
                    remaining_wait_time = for_total_time - elapsed_time
                else:
                    remaining_wait_time = 0
                if remaining_wait_time > 0:
                    wait_time_per_step = remaining_wait_time / (num_steps - 3)
                    time.sleep(wait_time_per_step)
                tien = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTotalAmount"]')))
                tien_number = tien.get_attribute('text')
                wrong = '\xa0'
                ma_kh[count_kh] = ma_kh[count_kh] + \
                    f": {tien_number.replace(f'{wrong}', ' ')}"
                error = ma_kh
                elapsed_time = time.time() - start_time
                if elapsed_time < for_total_time:
                    remaining_wait_time = for_total_time - elapsed_time
                else:
                    remaining_wait_time = 0

                if remaining_wait_time > 0:
                    wait_time_per_step = remaining_wait_time / (num_steps - 4)
                    time.sleep(wait_time_per_step)
                wait = WebDriverWait(driver, 10)
                tiep = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTitleAction"]')))
                tiep.click()
                time.sleep(1.5)
                wait = WebDriverWait(driver, 10)
                try:
                    xac_nhan = wait.until(EC.presence_of_element_located(
                        (AppiumBy.ID, 'com.vnpay.vpbankonline:id/tvTitleAction')))
                    xac_nhan.click()
                except:
                    try:
                        xac_nhan = wait.until(EC.presence_of_element_located(
                            (AppiumBy.ID, 'com.vnpay.vpbankonline:id/tvTitleAction')))
                        xac_nhan.click()
                    except:
                        pass
                    try:
                        xac_nhan = wait.until(EC.presence_of_element_located(
                            (AppiumBy.ID, 'com.vnpay.vpbankonline:id/tvTitleAction')))
                        xac_nhan.click()
                    except:
                        pass
                try:
                    new_bill = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/vMakeOtherPayment"]')))
                except:
                    pass
                if screenshot:
                    time_image = datetime.now().strftime("%H_%M_%S")
                    image_path = os.path.join(
                        os.getcwd(), "images", f"{device}_Bill_VPBank_{time_image}.png")
                    os.system(
                        f"adb -s {device} exec-out screencap -p > {image_path}")
                time.sleep(1.5)
                new_bill.click()
                driver.quit()

        def process_payment(func):
            with open('selected.json', 'r') as file:
                data = json.load(file)
            devices = data
            max_threads = len(devices)
            executor = ThreadPoolExecutor(max_workers=max_threads)
            futures = []
            for device in devices:
                futures.append(executor.submit(
                    func, device, account_name, ma_kh, screen_shot))
            for future in futures:
                future.result()
        process_payment(run)
        a = 0
        for i in error:
            df.loc[a, 'Thông Tin Hóa Đơn'] = i
            df.loc[a, 'Người Thanh Toán'] = self.userpayment
            a += 1
        time_excel = datetime.now().strftime("%H_%M")
        df.to_excel(f'vpbank{time_excel}.xlsx')

        def format_excel(file_path):
            df = pd.read_excel(file_path, engine='openpyxl')
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                        column[0].column)].width = adjusted_width
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
            wb.save(file_path)
        file_path = glob.glob('*.xlsx')
        for file_path in file_path:
            format_excel(file_path)

class Error(QDialog):
    def __init__(self, message):
        super().__init__()
        self.setWindowTitle("Lỗi")
        self.setWindowFlag(Qt.WindowCloseButtonHint)
        self.setWindowFlag(Qt.WindowMinimizeButtonHint)
        self.setFixedSize(200, 100)
        self.message = QLabel(message)
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.message)
        self.setLayout(self.layout)

        self.setStyleSheet("background-color: #282a36;")
        self.message.setStyleSheet("color: white; text-align: center; font-size: 20px;")


class ScanPopup(QDialog):
    def __init__(self):
        super().__init__()
        self.devices = get_devices()
        self.selected_devices = []

        self.setWindowTitle("Chọn thiết bị khởi chạy")
        self.layout = QVBoxLayout()
        self.setWindowFlags(self.windowFlags() | Qt.FramelessWindowHint)
        self.select_all_checkbox = QCheckBox("Chọn tất cả")
        self.select_all_checkbox.stateChanged.connect(self.select_all_devices)
        self.select_all_checkbox.setStyleSheet("border: none;")
        self.layout.addWidget(self.select_all_checkbox)
        self.empty_label = QLabel("Không tìm thấy thiết bị")
        self.empty_label.setStyleSheet(
            "font-size: 20px;border: none;text-align: center;")
        self.empty_label.hide()
        self.layout.addWidget(self.empty_label)
        self.draggable = True
        self.dragging_threshold = 5
        self.drag_start_position = None

        self.device_checkboxes = []
        for device_name in self.devices:
            checkbox = QCheckBox(device_name)
            checkbox.setStyleSheet("border: none;")
            self.device_checkboxes.append(checkbox)
            self.layout.addWidget(checkbox)

        self.button_layout = QHBoxLayout()

        if len(self.devices) > 0:
            self.save_button = QPushButton("Lưu")
            self.save_button.clicked.connect(self.save_devices)
            self.button_layout.addWidget(self.save_button)

        self.cancel_button = QPushButton("Huỷ")
        self.cancel_button.clicked.connect(self.cancel_scan)
        self.button_layout.addWidget(self.cancel_button)

        self.layout.addLayout(self.button_layout)

        self.setLayout(self.layout)

        if len(self.devices) == 0:
            self.empty_label.show()
            self.select_all_checkbox.hide()
        if len(self.device_checkboxes) == 1:
            self.select_all_checkbox.hide()
            self.device_checkboxes[0].setChecked(True)
        self.setStyleSheet("""
        * {
            font-family: 'JetBrains Mono', monospace;
            margin: 1px;
        }
        QWidget {
            background-color: #282a36;
            color: #f8f8f2;
            font-size: 16px;
            width: 300px;
            border-radius: 20px;
            border: 2px solid #44475A;
        }
        QCheckBox {
            spacing: 5px;
        }
        QCheckBox::indicator {
            width: 30px;
            height: 30px;
            border-radius: 15px;
        }
        QCheckBox::indicator:unchecked {
            background-color: #44475a;
            border: 1px solid #bd93f9;
            border-radius: 15px;
        }
        QCheckBox::indicator:checked {
            background-color: #bd93f9;
            border-radius: 15px;
        }
        QPushButton {
            background-color: #6272a4;
            border: none;
            color: #f8f8f2;
            padding: 15px;
            text-align: center;
            text-decoration: none;
            font-size: 16px;
            margin: 4px 2px;
        }
        QPushButton:hover {
            background-color: #44475a;
        }
    """)

    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton and self.draggable:
            self.drag_start_position = event.globalPos()

    def mouseMoveEvent(self, event: QMouseEvent):
        if self.drag_start_position is not None and self.draggable:
            if (event.globalPos() - self.drag_start_position).manhattanLength() >= self.dragging_threshold:
                self.move(self.pos() + event.globalPos() -
                          self.drag_start_position)
                self.drag_start_position = event.globalPos()

    def mouseReleaseEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton and self.draggable:
            self.drag_start_position = None

    def select_all_devices(self, state):
        for checkbox in self.device_checkboxes:
            checkbox.setChecked(state == Qt.Checked)

    def save_devices(self):
        selected_device_names = [
            checkbox.text() for checkbox in self.device_checkboxes if checkbox.isChecked()]
        selected_device_adb_names = select_devices(
            selected_device_names)
        if isinstance(selected_device_adb_names, str):
            QMessageBox.warning(self, "Lỗi", selected_device_adb_names)
        else:
            with open("selected.json", "w") as file:
                json.dump(selected_device_adb_names, file)
        self.close()

    def cancel_scan(self):
        self.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    subprocess.Popen(
        ["cmd.exe", "/c", "start", "appium", "--session-override"])
    window.show()
    sys.exit(app.exec_())
