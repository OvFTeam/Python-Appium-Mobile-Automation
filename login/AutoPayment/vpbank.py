import glob
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from telnetlib import EC

import openpyxl
import pandas as pd
from appium import webdriver
from appium.options.common.base import AppiumOptions
from appium.webdriver.common.appiumby import AppiumBy
from openpyxl.styles import Alignment
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# subprocess.Popen(["cmd.exe", "/c", "start", "appium", "--session-override"])


def vpbank(username="", ma_khach_hang=[], userpayment="", screenshot=0):
    df = pd.DataFrame()
    SCREENSHOT_PATH = os.path.join(os.getcwd(), "images")
    account_name = username
    ma_kh = ma_khach_hang
    screen_shot = False
    if screenshot == 2:
        screen_shot = True
    global error
    error = ["Không có lỗi"]

    def run(device, account_name, ma_kh, screenshot=False):
        global tien_number
        global error
        options = AppiumOptions()
        options.load_capabilities({
            "deviceName": f"{device}",
            "platformName": "Android",
            "automationName": "UiAutomator2",
            "udid": f"{device}"
        })
        try:
            driver = webdriver.Remote(
                "http://127.0.0.1:4723", options=options)
        except:
            driver = webdriver.Remote(
                "http://127.0.0.1:4723", options=options)
        wait = WebDriverWait(driver, 5)
        try:
            thanh_toan = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
            thanh_toan.click()
        except:
            pass
        global count_kh
        count_kh = 0
        kh = ma_kh[count_kh]
        total_wait_time = 0
        s_time = time.time()
        try:
            dien = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
            dien.click()
        except StaleElementReferenceException:
            dien = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '(//android.widget.ImageView[@resource-id=\"com.vnpay.vpbankonline:id/icon\"])[1]')))
            dien.click()
        except:
            pass
        end_time = time.time() - s_time
        for_total_time = total_wait_time - end_time
        for i in ma_kh:
            if i in error:
                continue
            num_steps = 5
            start_time = time.time()

            try:
                new_bill = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/ll_create_new_bill"]')))
                new_bill.click()
            except:
                pass
            nha_cungcap = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/llValue"]')))
            nha_cungcap.click()
            elapsed_time = time.time() - start_time
            if elapsed_time < for_total_time:
                remaining_wait_time = for_total_time - elapsed_time
            else:
                remaining_wait_time = 0

            if remaining_wait_time > 0:
                wait_time_per_step = remaining_wait_time / (num_steps - 1)
                time.sleep(wait_time_per_step)
            toan_quoc = wait.until(EC.presence_of_element_located((
                AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvName" and @text="Điện lực toàn quốc"]'
            )))
            toan_quoc.click()
            try:
                wait_tai_khoan_nguon = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='com.vnpay.vpbankonline:id/selectAccount']/android.widget.LinearLayout/android.widget.LinearLayout")))
                wait_tai_khoan_nguon.click()
            except StaleElementReferenceException:
                wait_tai_khoan_nguon = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='com.vnpay.vpbankonline:id/selectAccount']/android.widget.LinearLayout/android.widget.LinearLayout")))
                wait_tai_khoan_nguon.click()
            except:
                error = "Lỗi không xác định"
                return error

            def chon_tai_khoan_nguon(account):
                valid_input = False
                while not valid_input:
                    elements = driver.find_element(by=AppiumBy.XPATH, value="//android.widget.ListView[@resource-id='com.vnpay.vpbankonline:id/listAccount']") \
                        .find_elements(by=AppiumBy.CLASS_NAME, value="android.widget.TextView")
                    for element in elements:
                        if element.text == account:
                            valid_input = True
                            return element
                    return "Chỉ cần nhập số thẻ hoặc loại thẻ"

            chon_tai_khoan_nguon(account_name).click()
            elapsed_time = time.time() - start_time
            if elapsed_time < for_total_time:
                remaining_wait_time = for_total_time - elapsed_time
            else:
                remaining_wait_time = 0

            if remaining_wait_time > 0:
                wait_time_per_step = remaining_wait_time / (num_steps - 2)
                time.sleep(wait_time_per_step)

            def input(kh):
                global count_kh
                input_mkh = wait.until(EC.presence_of_element_located((
                    AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.vnpay.vpbankonline:id/edtCustomerCode"]')))
                input_mkh.send_keys(kh)
                tiep = wait.until(EC.presence_of_element_located(
                    (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTitleAction"]')))
                tiep.click()
                try:
                    error_text = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvError"]'))
                    ).get_attribute('text')
                    clear = wait.until(EC.presence_of_element_located(
                        (AppiumBy.XPATH, '//android.widget.ImageView[@resource-id="com.vnpay.vpbankonline:id/vClear"]')))
                    clear.click()
                    count_kh += 1
                    ma_kh[count_kh - 1] = kh + f': {error_text}'
                    if count_kh == len(ma_kh):
                        return error
                    kh = ma_kh[count_kh]
                    input(kh)
                except:
                    pass
                return ma_kh
            error = input(kh)

            elapsed_time = time.time() - start_time
            if elapsed_time < for_total_time:
                remaining_wait_time = for_total_time - elapsed_time
            else:
                remaining_wait_time = 0
            if remaining_wait_time > 0:
                wait_time_per_step = remaining_wait_time / (num_steps - 3)
                time.sleep(wait_time_per_step)
            tien = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTotalAmount"]')))
            tien_number = tien.get_attribute('text')
            wrong = '\xa0'
            ma_kh[count_kh] = ma_kh[count_kh] + \
                f": {tien_number.replace(f'{wrong}', ' ')}"
            error = ma_kh
            elapsed_time = time.time() - start_time
            if elapsed_time < for_total_time:
                remaining_wait_time = for_total_time - elapsed_time
            else:
                remaining_wait_time = 0

            if remaining_wait_time > 0:
                wait_time_per_step = remaining_wait_time / (num_steps - 4)
                time.sleep(wait_time_per_step)
            tiep = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTitleAction"]')))
            tiep.click()
            xac_nhan = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.TextView[@resource-id="com.vnpay.vpbankonline:id/tvTitleAction"]')))
            xac_nhan.click()
            wait = WebDriverWait(driver, 10)
            new_bill = wait.until(EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="com.vnpay.vpbankonline:id/vMakeOtherPayment"]')))
            if screenshot:
                time_image = datetime.now().strftime("%H_%M_%S")
                image_path = os.path.join(
                    os.getcwd(), "images", f"{device}_Bill_VPBank_{time_image}.png")
                os.system(
                    f"adb -s {device} exec-out screencap -p > {image_path}")
            time.sleep(1.5)
            new_bill.click()
            driver.quit()

    def process_payment(func):
        with open('selected.json', 'r') as file:
            data = json.load(file)
        devices = data
        max_threads = len(devices)
        executor = ThreadPoolExecutor(max_workers=max_threads)
        futures = []
        for device in devices:
            futures.append(executor.submit(
                func, device, account_name, ma_kh, screen_shot))
        for future in futures:
            future.result()
    process_payment(run)
    a = 0
    for i in error:
        df.loc[a, 'Thông Tin Hóa Đơn'] = i
        df.loc[a, 'Người Thanh Toán'] = userpayment
        a += 1
    time_excel = datetime.now().strftime("%H_%M")
    df.to_excel(f'vpbank{time_excel}.xlsx')

    def format_excel(file_path):
        df = pd.read_excel(file_path, engine='openpyxl')
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                    column[0].column)].width = adjusted_width
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
        wb.save(file_path)
    file_path = glob.glob('*.xlsx')
    for file_path in file_path:
        format_excel(file_path)

# VISA Shopee Platinum Credit asd ['PD29007350798', 'PD29007350798', 'PD29007350798'] 2
account_name = "VISA Shopee Platinum Credit"
userpayment = "asd"
ma_khach_hang = ['PD29007350798', 'PD29007350798', 'PD29007350798']
vpbank(account_name, ma_khach_hang, userpayment, 2)